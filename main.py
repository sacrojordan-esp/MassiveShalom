import requests, re, json
import pandas as pd
from urllib.parse import unquote
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Session setup
session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, text/plain, */*",
    "Origin": "https://cliente.shalom.pe",
    "Referer": "https://cliente.shalom.pe/",
    "X-Requested-With": "XMLHttpRequest"
})

# Cookies - se actualizan desde cookies.txt
session.cookies.set("shalomempresas_session", "eyJpdiI6IlwvZEs4VnVmdFJPVGFHSUVYQVFxUWp3PT0iLCJ2YWx1ZSI6IktGbnZpOElpXC84WEVIVzhlQ21Rbnlab0xnZDViMTQzdFZGb3RwN0lDZlYyUFVLamNtWDNNcVhqb0szWDFSU2xGIiwibWFjIjoiM2RjNjEyODZhYjc2N2UxMDc1Mzk1MmViZDU2NTU5MjNjYmY3YjljNmZkNTk2YTcyMzA4MzdjYTdjNDBhOGQ4YyJ9", domain="cliente.shalom.pe")
session.cookies.set("XSRF-TOKEN", "eyJpdiI6IjhxMFN4eHZzeU9wUURta2xqYmR4dWc9PSIsInZhbHVlIjoiNWF5bExNNG81bGxiWDN4VTNUN1BCZ056TXpzTUlcL2dNYUVtNzhKb2FoNVNVaVBpM1cxUmVKY2tFNG44SGdoTTAiLCJtYWMiOiI3NzEyNzRhOGIzZTY5Y2JkMzE4ODE4ODhhYzZkYmY1Yzc0MmYwM2IwOTFmYzk1NGJkZjUyZmM5ZGE1NTE5NzFiIn0%3D", domain="cliente.shalom.pe")

# Login page para CSRF
login_page = session.get("https://cliente.shalom.pe/")
csrf_token = re.search(r'name="csrf-token" content="(.*?)"', login_page.text)
csrf_token = csrf_token.group(1) if csrf_token else ""

# Headers
xsrf_token = unquote(session.cookies.get("XSRF-TOKEN", "") or "")
headers = {
    "X-Requested-With": "XMLHttpRequest",
    "X-XSRF-TOKEN": xsrf_token,
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json;charset=UTF-8",
    "Origin": "https://cliente.shalom.pe",
    "Referer": "https://cliente.shalom.pe/"
}

# Cargar destinos
print("Cargando destinos...")
r_destinos = session.get("https://cliente.shalom.pe/envia_ya/service_order/restricciones-categorias?origen=426", headers=headers)
destinos_raw = r_destinos.json()["data"]
# Normalizar nombres (quitar espacios extras)
destinos = {v["nombre_terminal"].upper().replace("  ", " ").strip(): int(k) for k, v in destinos_raw.items()}
print(f"Destinos: {len(destinos)}")

# Buscar destinatario por DNI
def buscar_destinatario(dni):
    dni_formatted = str(dni).zfill(8)

    try:
        r = session.post(
                "https://cliente.shalom.pe/envia_ya/person/search",
                json={
                    "documento": dni_formatted,
                    "type": "receiver"
                },
                headers=headers
            )

        print(f"    Status búsqueda: {r.status_code}")
        print(f"    Respuesta cruda: {r.text[:500]}")

        response_json = r.json()

        if not response_json.get("success"):
            print(f"    Búsqueda sin éxito para DNI {dni_formatted}: {response_json}")
            return None

        data_nested = response_json.get("data")
        print(f"    Data parseada: {data_nested}")

        if isinstance(data_nested, dict):
            posible_id = data_nested.get("id")
            print(f"    ID detectado en dict: {posible_id}")
            return posible_id

        print(f"    Formato inesperado en data para DNI {dni_formatted}: {data_nested}")
        return None

    except Exception as e:
        print(f"    Error buscando destinatario {dni_formatted}: {e}")
        return None

def parse_nested_json(value):
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value

# Calcular precio del envio
def calcular_precio(origen, destino_id):
    payload_calc = {"origin": origen, "destiny": destino_id, "width": "", "height": "", "length": "", "weight": ""}
    r = session.post("https://cliente.shalom.pe/envia_ya/tariff/calculate", json=payload_calc, headers=headers)
    data = r.json()
    if data.get("success"):
        return data.get("data", {}).get("price", 5)
    return 5

# Excel setup
df = pd.read_excel("MassiveShalom.xlsx")

# Filtrar filas vacias (sin DNI)
df = df[df["DNI"].notna()]

# Limpiar columna ESTADO
if "ESTADO" in df.columns:
    df["ESTADO"] = ""
else:
    df["ESTADO"] = ""
wb = load_workbook("MassiveShalom.xlsx")
ws = wb.active

verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
naranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Constantes
ORIGEN = 426
COSTO_FIJO = 5

# Funcion para formatear DNI con ceros a la izquierda
def formatear_dni(dni):
    # Convertir a string, quitar decimales si es float, y rellenar con ceros
    dni_str = str(int(float(dni)))
    return dni_str.zfill(8)

# Procesar cada fila
for idx, fila in df.iterrows():
    excel_row = idx + 2
    dni = formatear_dni(fila["DNI"])
    agencia = fila["AGENCIA SHALOM"]
    clave = str(fila["CLAVE"])
    
    print(f"Fila {idx + 1}: {fila['CLIENTE']}")
    
    # Buscar destino
    agencia_normalizada = str(agencia).upper().replace("  ", " ").strip()
    destino_id = destinos.get(agencia_normalizada)
    print(f"  Agencia: {agencia} -> ID: {destino_id}")
    if not destino_id:
        ws.cell(excel_row, 12, "NO PROCEDIO").fill = rojo
        continue
    
    # Buscar destinatario
    print(f"  Buscando DNI: {dni}")
    destinatario_id = buscar_destinatario(dni)
    print(f"  Destinatario ID: {destinatario_id}")
    if not destinatario_id:
        ws.cell(excel_row, 12, "NO PROCEDIO POR ID").fill = naranja
        continue
    
    # Calcular precio
    precio = calcular_precio(ORIGEN, destino_id)
    print(f"  Precio calculado: {precio}")
    
    # Enviar pedido
    payload = {
        "origen": ORIGEN, "destino": destino_id, "tipo_pago": "REMITENTE", "tipo_producto": 1090,
        "aereo": 0, "alto": "", "ancho": "", "cantidad": 1, "clave": clave, "contacto_doc": "",
        "costo": f"{precio}.00", "declaracion_jurada": "", "destinatario": dni, "destinatario_id": destinatario_id,
        "garantia": 0, "garantia_costo": 0, "garantia_monto": "0.00", "grrs": "[]", "largo": "", "peso": "",
        "remitente": "20612606103", "remitente_id": 1074, "servicio_cobranza": 0, "servicio_cobranza_costo": 0,
        "servicio_cobranza_datos": "{}", "tipo_producto_json": {"value": f"{precio}.00", "name": "Caja Paquete XS", "detalle": "15 x 20 x 12 cm"}
    }
    
    response = session.post("https://cliente.shalom.pe/envia_ya/service_order/save", json=payload, headers=headers)
    print(f"  Status: {response.status_code}, Response: {response.text[:100]}")
    
    try:
        resp_json = response.json()
        if resp_json.get("success"):
            data_resp = parse_nested_json(resp_json.get("data"))
            guia = data_resp.get("guia") if isinstance(data_resp, dict) else None
            print(f"  [OK] PROCEDIO - Guia: {guia}")
            ws.cell(excel_row, 12, "PROCEDIO").fill = verde
        else:
            msg = resp_json.get("message", "Error")
            print(f"  [ERROR] {msg}")
            ws.cell(excel_row, 12, f"NO PROCEDIO: {msg}").fill = rojo
    except Exception as e:
        print(f"  [ERROR] Respuesta invalida: {e}")
        ws.cell(excel_row, 12, "NO PROCEDIO").fill = rojo

wb.save("MassiveShalom.xlsx")
print("\nExcel guardado")
